import csv
import re
from collections import defaultdict
from copy import deepcopy
from datetime import time
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/Users/xayrullorozimatov/Desktop/HyperCodes/freelance/RMC")
WORKBOOK_PATH = Path("/Users/xayrullorozimatov/Documents/books/TEMURBEK SCHOOL.xlsx")
STUDENTS_CSV = ROOT / "docs" / "normalized" / "students_import.csv"
CLASSES_CSV = ROOT / "docs" / "normalized" / "classes_import.csv"
TEACHERS_CSV = ROOT / "docs" / "normalized" / "teachers_import.csv"


def compact(value):
    if value is None:
        return ""
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_name(value):
    text = compact(value)
    text = (
        text.replace("ʻ", "'")
        .replace("`", "'")
        .replace("‘", "'")
        .replace("’", "'")
        .replace("Gʻ", "G'")
        .replace("Joʻ", "Jo'")
    )
    text = re.sub(r"\s*\d+\s*$", "", text)
    text = re.sub(r"\.(?:\s*\d+.*)?$", "", text)
    text = re.sub(r"\s+", " ", text).strip(" .,")
    return text


def normalize_name_key(value):
    return normalize_name(value).upper()


def normalize_label_key(value):
    text = compact(value)
    text = (
        text.replace("ʻ", "'")
        .replace("`", "'")
        .replace("‘", "'")
        .replace("’", "'")
        .replace("Gʻ", "G'")
        .replace("Joʻ", "Jo'")
    )
    text = text.replace("'", "")
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def split_name(full_name):
    parts = normalize_name(full_name).split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[1:]), parts[0]


def normalize_phone(value):
    digits = re.sub(r"\D", "", compact(value))
    if not digits:
        return ""
    if len(digits) == 9:
        return f"+998{digits}"
    if len(digits) == 12 and digits.startswith("998"):
        return f"+{digits}"
    return f"+{digits}" if len(digits) >= 7 else ""


def read_csv(path):
    with path.open("r", encoding="utf-8", newline="") as fh:
        return list(csv.DictReader(fh))


def write_csv(path, rows, fieldnames):
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def looks_like_header(values):
    cells = [compact(value).lower() for value in values]
    return any(cell in ("t/r", "tr", "t / r", "n/r", "nr", "ismi familiyasi") for cell in cells) and any(
        any(token in cell for token in ("sana", "sinfi", "maktabi", "tel", "raqam")) for cell in cells
    )


def useful_title(values):
    cells = [compact(value) for value in values if compact(value)]
    if not cells:
        return ""
    first = cells[0]
    if normalize_label_key(first) in {"t/r", "tr", "t / r", "n/r", "nr", "ismi familiyasi"}:
        return ""
    if not re.search(r"[A-Za-z\u0400-\u04FF]", first):
        title = " ".join(cells)
        if not re.search(r"[A-Za-z\u0400-\u04FF]", title):
            return ""
        if len(cells) <= 3 and not re.match(r"^\d+\s", title):
            return title
        return ""
    later = cells[1:]
    if later and all(is_standalone_schedule(value) or is_numberish(value) or re.fullmatch(r"\d+(?:[.,]\d+)?%?", value) for value in later):
        return first
    title = " ".join(cells)
    if re.match(r"^\d+\s", title):
        return ""
    return title if len(cells) <= 3 else first


def is_schedule_text(value):
    text = compact(value).lower()
    return bool(
        re.search(r"\b(dush|du|sesh|se|chor|ch|pay|paysh|juma|ju|shanba|yak)\b", text)
        or re.search(r"\d{1,2}[:.]\d{2}|\b\d{3,4}\b(?:\s*[-_]\s*\b\d{3,4}\b)?", text)
    )


def is_standalone_schedule(value):
    text = compact(value).lower()
    if not text or not is_schedule_text(text):
        return False
    remainder = re.sub(r"\d{1,2}[:.]\d{2}(?:\s*[-_]\s*\d{1,2}[:.]\d{2})?", " ", text)
    remainder = re.sub(r"\b\d{3,4}\b", " ", remainder)
    remainder = re.sub(r"\b(dush|du|sesh|se|chor|ch|pay|paysh|juma|ju|shanba|yak|xona)\b", " ", remainder)
    remainder = re.sub(r"[(),/_-]", " ", remainder)
    remainder = re.sub(r"\s+", " ", remainder).strip()
    return remainder == ""


def is_numberish(value):
    return bool(re.fullmatch(r"\d+(?:\.0+)?", compact(value)))


def looks_like_note_row(values):
    cells = [compact(value) for value in values if compact(value)]
    if not cells:
        return False
    joined = " ".join(cells).lower()
    if any(token in joined for token in ("to'landi", "to'lansin", "grant", "ketgan", "berildi")):
        return True
    if re.search(r"=\s*\d", joined):
        return True
    return False


def teacher_from_sheet(sheet_name):
    name = re.sub(r"\bNew\b", "", sheet_name, flags=re.I)
    name = re.sub(r"\d+(?:[.,]\d+)?\s*%", "", name)
    name = re.sub(r"\b\d{4}\b", "", name)
    return re.sub(r"\s+", " ", name).strip()


def row_value(row, index):
    return compact(row[index]) if index < len(row) else ""


def extract_row_name(row, marker_index, name_index, seen_numbered):
    marker = row_value(row, marker_index)
    direct = row_value(row, name_index)
    shifted = row_value(row, name_index + 1)

    def has_letters(value):
        return bool(re.search(r"[A-Za-z\u0400-\u04FF]", value))

    if is_numberish(marker) and has_letters(direct):
        return direct, 0
    if is_numberish(marker) and is_numberish(direct) and has_letters(shifted):
        return shifted, 1
    if not marker and has_letters(direct):
        return direct, 0
    if seen_numbered and has_letters(marker):
        return marker, marker_index - name_index
    if seen_numbered and is_numberish(marker) and not direct and has_letters(shifted):
        return shifted, 1
    return "", 0


def extract_blocks():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    blocks = defaultdict(list)
    for ws in workbook.worksheets:
        if re.search(r"\(\d+\)\s*$", ws.title):
            continue
        if ws.title in {"Yangi Kelganlar", "Dars jadvali"}:
            continue
        teacher = teacher_from_sheet(ws.title)
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        previous_title = ""
        index = 0
        while index < len(rows):
            row = rows[index]
            title = useful_title(row)
            if title and not is_standalone_schedule(title):
                previous_title = title

            if not looks_like_header(row):
                index += 1
                continue

            cells = [compact(value) for value in row]
            marker_index = next(
                i for i, cell in enumerate(cells) if cell.lower() in ("t/r", "tr", "t / r", "n/r", "nr", "ismi familiyasi")
            )
            name_index = marker_index + 1
            header_name = cells[name_index] if name_index < len(cells) else ""
            class_name = (
                previous_title
                if is_standalone_schedule(header_name) or header_name.lower() in ("ismi familiyasi", "ism familiya", "fio", "full name", "name")
                else header_name or previous_title
            )
            class_name = compact(class_name)

            phone_index = next((i for i, cell in enumerate(cells) if "tel" in cell.lower() or "raqam" in cell.lower()), -1)
            school_class_index = next((i for i, cell in enumerate(cells) if "sinf" in cell.lower()), -1)
            school_name_index = next((i for i, cell in enumerate(cells) if "maktab" in cell.lower()), -1)

            block_rows = []
            seen_numbered = False
            index += 1
            while index < len(rows):
                current = rows[index]
                if looks_like_header(current):
                    break
                next_title = useful_title(current)
                nonempty = sum(1 for value in current if compact(value))
                if next_title and nonempty <= 3:
                    break

                marker = row_value(current, marker_index)
                name, offset = extract_row_name(current, marker_index, name_index, seen_numbered)
                if not name or looks_like_note_row(current):
                    index += 1
                    continue

                is_student_row = False
                if is_numberish(marker):
                    is_student_row = True
                    seen_numbered = True
                elif seen_numbered and re.search(r"[A-Za-z\u0400-\u04FF]", name) and nonempty <= 5:
                    is_student_row = True

                if is_student_row:
                    block_rows.append(
                        {
                            "name": normalize_name(name),
                            "phone": normalize_phone(current[phone_index + offset]) if phone_index >= 0 and phone_index + offset < len(current) else "",
                            "school_class": compact(current[school_class_index + offset]) if school_class_index >= 0 and school_class_index + offset < len(current) else "",
                            "school_name": compact(current[school_name_index + offset]) if school_name_index >= 0 and school_name_index + offset < len(current) else "",
                        }
                    )
                index += 1

            if block_rows and class_name:
                deduped = []
                seen = set()
                for row_data in block_rows:
                    key = (normalize_name_key(row_data["name"]), row_data["phone"])
                    if key in seen:
                        continue
                    seen.add(key)
                    deduped.append(row_data)
                blocks[teacher].append({"class_name": class_name, "rows": deduped})
    return blocks


def manual_block_rows(ws, class_name, start_row, end_row):
    rows = []
    seen = set()
    for row_index in range(start_row, end_row + 1):
        values = [ws.cell(row_index, column).value for column in range(1, 21)]
        c1 = compact(values[0])
        c2 = compact(values[1])
        c3 = compact(values[2])

        if is_numberish(c1) and re.search(r"[A-Za-z\u0400-\u04FF]", c2):
            name = c2
        elif is_numberish(c2) and re.search(r"[A-Za-z\u0400-\u04FF]", c3):
            name = c3
        elif not c1 and not c2 and re.search(r"[A-Za-z\u0400-\u04FF]", c3):
            name = c3
        elif re.search(r"[A-Za-z\u0400-\u04FF]", c1) and not is_standalone_schedule(c1):
            name = c1
        elif re.search(r"[A-Za-z\u0400-\u04FF]", c2) and not is_standalone_schedule(c2):
            name = c2
        else:
            continue

        name = normalize_name(name)
        if not name or normalize_label_key(name) in {"t/r", "n/r", "ismi familiyasi", "ismi familyasi"}:
            continue
        if any(token in normalize_label_key(name) for token in {"otganlar", "ketganlar", "yangi guruhlari"}):
            continue
        if normalize_label_key(name) in {"nafosat", "elzora", "umar"}:
            continue
        key = normalize_name_key(name)
        if key in seen:
            continue
        seen.add(key)
        rows.append({"name": name, "phone": "", "school_class": "", "school_name": ""})
    return {"class_name": class_name, "rows": rows}


def build_manual_blocks():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    manual_specs = {
        "Sodiqov Xurshid": [
            ("Sodiqov Xurshid 50%", "B2 intro", 6, 18),
            ("Sodiqov Xurshid 50%", "B1 intro", 32, 51),
            ("Sodiqov Xurshid 50%", "A2 Flyers", 58, 77),
            ("Sodiqov Xurshid 50%", "fun for movers", 87, 99),
        ],
        "Shaxobov Temurbek": [
            ("Shaxobov Temurbek 2026 50 %", "B2 Multilevel September", 3, 37),
            ("Shaxobov Temurbek 2026 50 %", "Starter New", 48, 80),
            ("Shaxobov Temurbek 2026 50 %", "B1 Intro", 92, 131),
            ("Shaxobov Temurbek 2026 50 %", "Special Grammar B1 Plus", 134, 175),
            ("Shaxobov Temurbek 2026 50 %", "IELTS 7.0 Premier", 178, 220),
            ("Shaxobov Temurbek 2026 50 %", "A1 Movers KIDS", 223, 256),
            ("Shaxobov Temurbek 2026 50 %", "B1 plus Morning", 267, 296),
            ("Shaxobov Temurbek 2026 50 %", "A2 Plus Morning", 308, 342),
        ],
        "Oktamov Aminjon": [
            ("Oktamov Aminjon 50%", "Matematika 10", 45, 68),
            ("Oktamov Aminjon 50%", "Matematika 3 16:00", 80, 104),
            ("Oktamov Aminjon 50%", "Matematika Individual 17:30", 108, 126),
            ("Oktamov Aminjon 50%", "Matematika 2", 130, 157),
        ],
        "Meliqulov Ibrohim": [
            ("Meliqulov Ibrohim 50%", "Arab tili 1.0 New", 6, 42),
            ("Meliqulov Ibrohim 50%", "Arab tili 3.0 New", 77, 112),
            ("Meliqulov Ibrohim 50%", "Arab tili 4.0 New", 116, 146),
            ("Meliqulov Ibrohim 50%", "Arab tili 5.0 New", 149, 176),
            ("Meliqulov Ibrohim 50%", "Arab tili 2", 200, 219),
            ("Meliqulov Ibrohim 50%", "Arab tili 1", 225, 259),
        ],
        "Fazliyev Ramziddin": [
            ("Fazliyev Ramziddin 45%", "Pirizident maktab Ramziddin ustoz", 6, 9),
        ],
    }

    manual_blocks = defaultdict(list)
    for teacher_name, specs in manual_specs.items():
        for sheet_name, class_name, start_row, end_row in specs:
            manual_blocks[teacher_name].append(manual_block_rows(workbook[sheet_name], class_name, start_row, end_row))
    return manual_blocks


def next_enrollment_number(rows):
    max_number = 0
    for row in rows:
        match = re.search(r"(\d+)$", row.get("enrollment_number", ""))
        if match:
            max_number = max(max_number, int(match.group(1)))
    current = max_number + 1
    while True:
        yield f"STU-{current:05d}"
        current += 1


def next_username(existing_rows):
    used = {row.get("username", "").strip() for row in existing_rows if row.get("username", "").strip()}

    def create(full_name):
        first_name, _ = split_name(full_name)
        base = re.sub(r"[^a-z0-9]+", "", normalize_name(first_name).lower()) or "student"
        candidate = base
        counter = 2
        while candidate in used:
            candidate = f"{base}{counter}"
            counter += 1
        used.add(candidate)
        return candidate

    return create


def get_teacher_and_class_maps(teachers, classes):
    teacher_by_name = {}
    for teacher in teachers:
        full_name = f"{teacher['last_name']} {teacher['first_name']}".strip()
        teacher_by_name[normalize_label_key(full_name)] = teacher["employee_id"]

    class_buckets = defaultdict(list)
    for class_row in classes:
        class_buckets[(class_row["teacher_employee_id"], class_row["class_name"])].append(class_row)

    for bucket in class_buckets.values():
        bucket.sort(key=lambda row: row["class_code"])

    return teacher_by_name, class_buckets


def find_block(blocks, teacher_name, class_name, occurrence=0):
    teacher_key = normalize_label_key(teacher_name)
    resolved_teacher = next((name for name in blocks if normalize_label_key(name) == teacher_key), teacher_name)
    target_key = normalize_label_key(class_name)
    matches = [block for block in blocks[resolved_teacher] if normalize_label_key(block["class_name"]) == target_key]
    if not matches:
        matches = [
            block
            for block in blocks[resolved_teacher]
            if target_key in normalize_label_key(block["class_name"]) or normalize_label_key(block["class_name"]) in target_key
        ]
    if occurrence >= len(matches):
        raise KeyError(f"Missing workbook block: {teacher_name} / {class_name} / {occurrence}")
    return matches[occurrence]


def roster_by_class_code(students):
    roster = defaultdict(list)
    for row in students:
        roster[row["class_code"]].append(row)
    return roster


def student_index(students):
    by_name = defaultdict(list)
    by_name_phone = defaultdict(list)
    for row in students:
        full_name = f"{row.get('last_name', '')} {row.get('first_name', '')}".strip()
        key = normalize_name_key(full_name)
        by_name[key].append(row)
        by_name_phone[(key, normalize_phone(row.get("phone", "")))].append(row)
    return by_name, by_name_phone


def assign_roster(
    students,
    teacher_employee_id,
    class_row,
    desired_rows,
    enrollments,
    make_username,
):
    target_code = class_row["class_code"]
    target_name = class_row["class_name"]
    class_row["capacity"] = str(len(desired_rows))
    current_by_code = roster_by_class_code(students)
    existing_name_index, existing_name_phone_index = student_index(students)

    target_current = current_by_code.get(target_code, [])
    desired_keys = [normalize_name_key(row["name"]) for row in desired_rows]
    desired_key_set = set(desired_keys)

    # Remove students no longer in the target roster from this class.
    for row in list(target_current):
        row_key = normalize_name_key(f"{row['last_name']} {row['first_name']}")
        if row_key not in desired_key_set:
            students.remove(row)

    # Refresh indexes after removals.
    current_by_code = roster_by_class_code(students)
    existing_name_index, existing_name_phone_index = student_index(students)
    target_current = current_by_code.get(target_code, [])

    for desired in desired_rows:
        name_key = normalize_name_key(desired["name"])
        phone = desired["phone"]
        row = None
        target_matches = [
            current_row
            for current_row in target_current
            if normalize_name_key(f"{current_row['last_name']} {current_row['first_name']}") == name_key
        ]
        if phone:
            target_matches = [
                current_row
                for current_row in target_matches
                if normalize_phone(current_row.get("phone", "")) in {"", phone}
            ] or target_matches
        if target_matches:
            row = target_matches[0]
        if row is None:
            template = None
            if phone and existing_name_phone_index.get((name_key, phone)):
                template = existing_name_phone_index[(name_key, phone)][0]
            elif existing_name_index.get(name_key):
                template = existing_name_index[name_key][0]
            first_name, last_name = split_name(desired["name"])
            row = {
                "enrollment_number": next(enrollments),
                "first_name": template.get("first_name", first_name) if template else first_name,
                "last_name": template.get("last_name", last_name) if template else last_name,
                "email": template.get("email", "") if template else "",
                "phone": phone,
                "date_of_birth": template.get("date_of_birth", "") if template else "",
                "parent_name": template.get("parent_name", "") if template else "",
                "parent_phone": template.get("parent_phone", "") if template else "",
                "gender": template.get("gender", "") if template else "",
                "status": "Active",
                "username": make_username(desired["name"]),
                "password": "012345678",
                "teacher_employee_id": teacher_employee_id,
                "class_name": target_name,
                "class_code": target_code,
                "school_name": desired["school_name"] or (template.get("school_name", "") if template else ""),
                "school_class": desired["school_class"] or (template.get("school_class", "") if template else ""),
            }
            students.append(row)
        else:
            row["teacher_employee_id"] = teacher_employee_id
            row["class_name"] = target_name
            row["class_code"] = target_code
            if phone:
                row["phone"] = phone
            if desired["school_name"]:
                row["school_name"] = desired["school_name"]
            if desired["school_class"]:
                row["school_class"] = desired["school_class"]
            row["status"] = row.get("status") or "Active"
            row["password"] = row.get("password") or "012345678"
            row["username"] = row.get("username") or make_username(desired["name"])
        current_by_code = roster_by_class_code(students)
        target_current = current_by_code.get(target_code, [])

    seen_target_names = set()
    for row in list(students):
        if row["class_code"] != target_code:
            continue
        row_key = normalize_name_key(f"{row['last_name']} {row['first_name']}")
        if row_key in seen_target_names:
            students.remove(row)
            continue
        seen_target_names.add(row_key)


def rename_class(classes, students, class_code, new_name):
    for class_row in classes:
        if class_row["class_code"] == class_code:
            class_row["class_name"] = new_name
    for student in students:
        if student["class_code"] == class_code:
            student["class_name"] = new_name


def remove_class(classes, students, class_code):
    classes[:] = [class_row for class_row in classes if class_row["class_code"] != class_code]
    students[:] = [student for student in students if student["class_code"] != class_code]


def main():
    students = read_csv(STUDENTS_CSV)
    classes = read_csv(CLASSES_CSV)
    teachers = read_csv(TEACHERS_CSV)
    workbook_blocks = extract_blocks()
    for teacher_name, manual_blocks in build_manual_blocks().items():
        existing_keys = {normalize_label_key(block["class_name"]) for block in manual_blocks}
        workbook_blocks[teacher_name] = [
            block for block in workbook_blocks[teacher_name] if normalize_label_key(block["class_name"]) not in existing_keys
        ] + manual_blocks
    teacher_by_name, class_buckets = get_teacher_and_class_maps(teachers, classes)

    enrollments = next_enrollment_number(students)
    make_username = next_username(students)

    def class_row(teacher_name, class_name, occurrence=0):
        teacher_id = teacher_by_name[normalize_label_key(teacher_name)]
        bucket = next(
            (
                rows
                for (bucket_teacher_id, bucket_name), rows in class_buckets.items()
                if bucket_teacher_id == teacher_id and normalize_label_key(bucket_name) == normalize_label_key(class_name)
            ),
            [],
        )
        if occurrence >= len(bucket):
            raise KeyError(f"Missing class row: {teacher_name} / {class_name} / {occurrence}")
        return teacher_id, bucket[occurrence]

    # Xurshid Sodiqov fixes.
    for class_name in ["B2 intro", "B1 intro", "A2 Flyers", "fun for movers"]:
        block = find_block(workbook_blocks, "Sodiqov Xurshid", class_name)
        current_name = "B1 intro (Xurshid)" if class_name == "B1 intro" else class_name
        teacher_id, row = class_row("Sodiqov Xurshid", current_name)
        assign_roster(students, teacher_id, row, block["rows"], enrollments, make_username)

    # Temurbek Shaxobov targeted fixes.
    for teacher_name, current_name, workbook_name in [
        ("Shaxobov Temurbek", "B2 Multilevel September", "B2 Multilevel September"),
        ("Shaxobov Temurbek", "Starter New", "Starter New"),
        ("Shaxobov Temurbek", "B1 Intro (Temurbek)", "B1 Intro"),
        ("Shaxobov Temurbek", "Special Grammar B1 Plus", "Special Grammar B1 Plus"),
        ("Shaxobov Temurbek", "IELTS 7.0 Premier", "IELTS 7.0 Premier"),
        ("Shaxobov Temurbek", "B1 plus Morning", "B1 plus Morning"),
        ("Shaxobov Temurbek", "A2 Plus Morning", "A2 Plus Morning"),
        ("Shaxobov Temurbek", "A1 Movers KIDS", "A1 Movers KIDS"),
    ]:
        occurrence = 0
        block = find_block(workbook_blocks, teacher_name, workbook_name, occurrence)
        teacher_id, row = class_row(teacher_name, current_name)
        assign_roster(students, teacher_id, row, block["rows"], enrollments, make_username)

    # Javhar: A2 Plus and misnamed B1 intro class.
    block = find_block(workbook_blocks, "Shaxobov Javhar", "A2 Plus morning 9:30")
    teacher_id, row = class_row("Shaxobov Javhar", "A2 Plus morning 9:30")
    assign_roster(students, teacher_id, row, block["rows"], enrollments, make_username)

    try:
        teacher_id, afternoon_row = class_row("Shaxobov Javhar", "A2 flyers Afternoon (Javhar Ustoz)")
    except KeyError:
        teacher_id, afternoon_row = class_row("Shaxobov Javhar", "B1 INTRO Afternoon (Javhar Ustoz)")
    rename_class(classes, students, afternoon_row["class_code"], "B1 INTRO Afternoon (Javhar Ustoz)")
    afternoon_row["class_name"] = "B1 INTRO Afternoon (Javhar Ustoz)"
    block = find_block(workbook_blocks, "Shaxobov Javhar", "Sesh, Pay, Shan 13:30")
    assign_roster(students, teacher_id, afternoon_row, block["rows"], enrollments, make_username)

    # Amonov and Tursunov counts.
    block = find_block(workbook_blocks, "Amonov Erkin", "A2 flyers")
    teacher_id, row = class_row("Amonov Erkin", "A2 flyers")
    assign_roster(students, teacher_id, row, block["rows"], enrollments, make_username)
    remove_class(classes, students, "TCH-004-STARTER")
    remove_class(classes, students, "TCH-004-STARTER-2")

    for class_name in [
        "A2 Plus (Islomjon ustoz)",
        "Fun For movers (Islomjon ustoz)",
        "Individual (Islomjon ustoz)",
        "Starter (Islomjon ustoz)",
    ]:
        block = find_block(workbook_blocks, "Tursunov Islomjon", class_name)
        teacher_id, row = class_row("Tursunov Islomjon", class_name)
        assign_roster(students, teacher_id, row, block["rows"], enrollments, make_username)
    afternoon_class_code = "TCH-019-STARTER-AFTERNOON-ISLOMJON-USTOZ"
    if not any(class_row["class_code"] == afternoon_class_code for class_row in classes):
        classes.append(
            {
                "class_name": "starter (Afternoon)",
                "class_code": afternoon_class_code,
                "teacher_employee_id": teacher_id,
                "level": "",
                "section": '{"days":["Monday","Wednesday","Friday"],"time":"15:00","endTime":"16:30"}',
                "capacity": "",
                "room_number": "",
                "start_date": "",
                "end_date": "",
                "payment_amount": "",
                "payment_frequency": "Monthly",
            }
        )
        class_buckets[(teacher_id, "starter (Afternoon)")].append(classes[-1])
    afternoon_block = find_block(workbook_blocks, "Tursunov Islomjon", "starter (Afternoon)")
    _, afternoon_row = class_row("Tursunov Islomjon", "starter (Afternoon)")
    assign_roster(students, teacher_id, afternoon_row, afternoon_block["rows"], enrollments, make_username)

    # O'ktamov Aminjon reconciliations.
    for class_name in ["Matematika 10", "Matematika Individual 17:30"]:
        block = find_block(workbook_blocks, "O'ktamov Aminjon", class_name)
        teacher_id, row = class_row("O'ktamov Aminjon", class_name)
        assign_roster(students, teacher_id, row, block["rows"], enrollments, make_username)

    # Use the first workbook Matematika 2 block (28 students) as the main class roster.
    block = find_block(workbook_blocks, "O'ktamov Aminjon", "Matematika 2", 0)
    teacher_id, row = class_row("O'ktamov Aminjon", "Matematika 2")
    assign_roster(students, teacher_id, row, block["rows"], enrollments, make_username)

    # Meliqulov explicit new blocks.
    for class_name in ["Arab tili 1.0 New", "Arab tili 3.0 New", "Arab tili 4.0 New", "Arab tili 5.0 New", "Arab tili 2", "Arab tili 1"]:
        block = find_block(workbook_blocks, "Meliqulov Ibrohim", class_name)
        teacher_id, row = class_row("Meliqulov Ibrohim", class_name.replace("1.0", "1.0").replace("3.0", "3.0").replace("4.0", "4.0").replace("5.0", "5.0"))
        assign_roster(students, teacher_id, row, block["rows"], enrollments, make_username)

    # Split or refresh timing-based kids groups.
    for teacher_name, class_name, workbook_name, occurrence in [
        ("Bahtiyorova Laylo", "Kids Laylo", "Dush,Cor,Juma 13:30", 0),
        ("Bahtiyorova Laylo", "Kids Laylo", "Dush,Cor,Juma 14:30", 1),
        ("Mamasolixova Zubayda", "Kids Zubayda", "Kids Zubayda", 0),
        ("Mamasolixova Zubayda", "Kids Zubayda", "Kids Zubayda", 1),
        ("Ahatova Farangiz", "Kids Farangiz", "Kids Farangiz", 0),
        ("Ahatova Farangiz", "Kids Farangiz", "Kids Farangiz", 1),
    ]:
        block_occurrence = occurrence if workbook_name in {"Kids Zubayda", "Kids Farangiz"} else 0
        block = find_block(workbook_blocks, teacher_name, workbook_name, block_occurrence)
        teacher_id, row = class_row(teacher_name, class_name, occurrence)
        assign_roster(students, teacher_id, row, block["rows"], enrollments, make_username)

    # Rus tili split by existing duplicate class codes.
    for occurrence, workbook_name in enumerate(["Sesh Pay Shanba 13:30 giramatika", "Sesh Pay Shan rus tili"]):
        block = find_block(workbook_blocks, "RUSTili", workbook_name)
        teacher_id, row = class_row("RUSTili", "Rus tili Afternoon Shoxsanam", occurrence)
        assign_roster(students, teacher_id, row, block["rows"], enrollments, make_username)

    # Fazliyev Ramziddin targeted roster correction.
    block = find_block(workbook_blocks, "Fazliyev Ramziddin", "Pirizident maktab Ramziddin ustoz")
    teacher_id, row = class_row("Fazliyev Ramziddin", "Pirizident maktab Ramziddin ustoz")
    assign_roster(students, teacher_id, row, block["rows"], enrollments, make_username)

    # Shamsiya: keep afternoon group and add/refresh the morning group using the spare class code.
    teacher_id, row = class_row("Xursandmurodova Shamsiya", "Kids English (Shamsiyabegim)")
    block = find_block(workbook_blocks, "Xursandmurodova Shamsiya", "Sesh pay shan 14:00")
    assign_roster(students, teacher_id, row, block["rows"], enrollments, make_username)

    # Add a new class for Kids English morning by repurposing the existing Shamsiyabegim name to the morning label.
    # Preserve the original afternoon code above and use a brand new code for the morning group.
    new_class_code = "TCH-021-KIDS-ENGLISH-MORNING"
    if not any(class_row["class_code"] == new_class_code for class_row in classes):
        classes.append(
            {
                "class_name": "Kids English morning",
                "class_code": new_class_code,
                "teacher_employee_id": teacher_id,
                "level": "",
                "section": '{"days":["Monday","Wednesday","Friday"],"time":"10:30","endTime":"12:00"}',
                "capacity": "",
                "room_number": "",
                "start_date": "",
                "end_date": "",
                "payment_amount": "",
                "payment_frequency": "Monthly",
            }
        )
        class_buckets[(teacher_id, "Kids English morning")].append(classes[-1])
    morning_block = find_block(workbook_blocks, "Xursandmurodova Shamsiya", "Kids English morning")
    _, morning_row = class_row("Xursandmurodova Shamsiya", "Kids English morning")
    assign_roster(students, teacher_id, morning_row, morning_block["rows"], enrollments, make_username)

    # Re-sort students by enrollment number for stable output.
    def student_sort_key(row):
        match = re.search(r"(\d+)$", row.get("enrollment_number", ""))
        return int(match.group(1)) if match else 10**9

    actual_counts = defaultdict(int)
    for student in students:
        actual_counts[student["class_code"]] += 1
    for class_row in classes:
        if class_row["class_code"] in actual_counts:
            class_row["capacity"] = str(actual_counts[class_row["class_code"]])

    students.sort(key=student_sort_key)

    write_csv(STUDENTS_CSV, students, list(students[0].keys()))
    write_csv(CLASSES_CSV, classes, list(classes[0].keys()))


if __name__ == "__main__":
    main()
